from openpyxl import Workbook

# สร้าง Workbook object
wb = Workbook()

# สร้าง Sheet 1
# ws1 = wb.active
# ws1.title = "Sheet 1"

# สร้าง Sheet 2
ws2 = wb.create_sheet("Sheet 2")

# สร้าง Sheet 3 และใส่เป็น index ที่ 0 (เป็น Sheet แรก)
ws3 = wb.create_sheet("Sheet 3", 0)

# พิมพ์ข้อมูลลงใน Sheet 1
# ws1['A1'] = 'Hello World!'

# พิมพ์ข้อมูลลงใน Sheet 2
ws2['A1'] = 'This is Sheet 2'

# พิมพ์ข้อมูลลงใน Sheet 3
ws3['A1'] = 'This is Sheet 3'

for i in range(2):
    ws = wb.create_sheet(str(i))
# บันทึก Workbook
wb.save("my_workbook.xlsx")
